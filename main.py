import os
import io
import json
import httpx
import boto3
import pdfplumber
from docx import Document
from openpyxl import load_workbook
from fastapi import FastAPI, Form
from fastapi.responses import StreamingResponse
import asyncio
from concurrent.futures import ThreadPoolExecutor

# -------------------------
# CONFIG
# -------------------------

MINIO_URL = os.getenv("MINIO_URL")
MINIO_ACCESS = os.getenv("MINIO_ACCESS_KEY")
MINIO_SECRET = os.getenv("MINIO_SECRET_KEY")
OLLAMA_URL = os.getenv("OLLAMA_URL").rstrip("/")
MODEL_NAME = "llama3"
BUCKET = "llm-pdfs"
DOCUMENTS_REFRESH_INTERVAL = 60  # segundos
CHUNK_SIZE = 3000
MAX_CHUNKS_PER_PROMPT = 50  # limitar número de chunks enviados ao LLM

app = FastAPI(title="HELPER BACKEND")

s3 = boto3.client(
    "s3",
    endpoint_url=MINIO_URL,
    aws_access_key_id=MINIO_ACCESS,
    aws_secret_access_key=MINIO_SECRET
)

DOCUMENTS = []
executor = ThreadPoolExecutor(max_workers=8)

# -------------------------
# EXTRAÇÃO DE DOCUMENTOS
# -------------------------

def extract_pdf(b):
    text = ""
    with pdfplumber.open(io.BytesIO(b)) as pdf:
        for p in pdf.pages:
            t = p.extract_text()
            if t:
                text += t + "\n"
    return text

def extract_docx(b):
    doc = Document(io.BytesIO(b))
    return "\n".join(p.text for p in doc.paragraphs)

def extract_excel(b):
    text = ""
    wb = load_workbook(io.BytesIO(b), data_only=True)
    for sheet in wb.worksheets:
        for row in sheet.iter_rows(values_only=True):
            row_text = " | ".join(str(cell) for cell in row if cell is not None)
            if row_text:
                text += row_text + "\n"
    return text

def chunk_text(text, size=CHUNK_SIZE):
    return [text[i:i+size] for i in range(0, len(text), size)]

# -------------------------
# LOAD DOCUMENTS
# -------------------------

async def load_document(key, raw):
    if key.endswith(".pdf"):
        text = await asyncio.get_event_loop().run_in_executor(executor, extract_pdf, raw)
    elif key.endswith(".docx"):
        text = await asyncio.get_event_loop().run_in_executor(executor, extract_docx, raw)
    elif key.endswith(".xlsx") or key.endswith(".xls"):
        text = await asyncio.get_event_loop().run_in_executor(executor, extract_excel, raw)
    else:
        text = raw.decode("utf-8", errors="ignore")

    chunks = chunk_text(text)
    return [{"filename": key, "text": c} for c in chunks]

async def load_documents():
    global DOCUMENTS
    print("[INFO] Carregando documentos do MinIO...")
    resp = s3.list_objects_v2(Bucket=BUCKET)
    if "Contents" not in resp:
        print("[INFO] Nenhum documento encontrado.")
        return

    tasks = []
    for item in resp["Contents"]:
        key = item["Key"]
        obj = s3.get_object(Bucket=BUCKET, Key=key)
        raw = obj["Body"].read()
        tasks.append(load_document(key, raw))

    results = await asyncio.gather(*tasks)
    DOCUMENTS = [chunk for doc in results for chunk in doc]
    print(f"[INFO] Total chunks carregados: {len(DOCUMENTS)}")

async def documents_refresher():
    while True:
        await load_documents()
        await asyncio.sleep(DOCUMENTS_REFRESH_INTERVAL)

# -------------------------
# CONTEXTO SELECIONADO
# -------------------------

def get_relevant_context():
    # envia apenas os últimos N chunks para o LLM
    return "\n\n".join(
        f"Arquivo: {doc['filename']}\n{doc['text']}"
        for doc in DOCUMENTS[-MAX_CHUNKS_PER_PROMPT:]
    )

# -------------------------
# STREAM DO LLM
# -------------------------

async def stream_llm(prompt, context):
    url = f"{OLLAMA_URL}/api/generate"

    system_prompt = """
Você é o HELPER — a inteligência artificial do Rafael.

Regras IMPORTANTES:
- Use SOMENTE o contexto fornecido.
- Não invente NADA.
- Não responda com coisas que não vieram dos documentos.
- Se a resposta não estiver nos arquivos, diga EXATAMENTE:
  "Não encontrei essa informação nos documentos da base de conhecimento."

Diretrizes de Resposta:
- Cite o nome do arquivo quando fizer referência.
- Seja direto, organizado e objetivo.
- Nunca gere texto fora do contexto permitido.
- Não gere suposições ou inferências.
"""

    final_prompt = f"""
{system_prompt}

### CONTEXTO
{context}

### PERGUNTA DO USUÁRIO
{prompt}

### RESPOSTA BASEADA SOMENTE NO CONTEXTO:
"""

    payload = {
        "model": MODEL_NAME,
        "prompt": final_prompt,
        "stream": True,
        "options": {
            "temperature": 0,
            "num_ctx": 4096,
            "num_predict": 512,
            "top_p": 0.8,
            "repeat_penalty": 1.1
        }
    }

    async with httpx.AsyncClient(timeout=None) as client:
        async with client.stream("POST", url, json=payload) as resp:
            async for line in resp.aiter_lines():
                if not line:
                    continue
                try:
                    data = json.loads(line)
                    if "response" in data:
                        yield data["response"]
                except:
                    pass

# -------------------------
# WARMUP
# -------------------------

async def warmup():
    print("[INFO] Warmup do modelo iniciando...")
    context = get_relevant_context()
    prompt = "como criar uma fila rabbitmq?"
    async for _ in stream_llm(prompt, context):
        pass
    print("[INFO] Warmup concluído!")

# -------------------------
# ENDPOINTS
# -------------------------

@app.on_event("startup")
async def startup_event():
    await load_documents()
    asyncio.create_task(documents_refresher())
    await warmup()

@app.post("/stream")
async def stream(prompt: str = Form(...)):
    context = get_relevant_context()

    async def token_stream():
        async for token in stream_llm(prompt, context):
            yield token

    return StreamingResponse(token_stream(), media_type="text/plain")

